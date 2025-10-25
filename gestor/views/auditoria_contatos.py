# gestor/views/auditoria_contatos.py

from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.db.models import Q
from django.core.paginator import Paginator
from core.models import ChatwootLastConversation, Cliente
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime

@login_required
def auditoria_contatos_list(request):
    """Lista conversas do Chatwoot com auditoria de contatos"""

    # Busca
    busca = request.GET.get('q', '')

    # Query base
    conversas = ChatwootLastConversation.objects.all().order_by('-last_activity')

    # Aplicar busca
    if busca:
        conversas = conversas.filter(
            Q(name__icontains=busca) |
            Q(email__icontains=busca) |
            Q(phone__icontains=busca) |
            Q(codigo__icontains=busca) |
            Q(tipo__icontains=busca)
        )

    # PAGINAÇÃO PRIMEIRO (antes de buscar clientes)
    total_conversas = conversas.count()
    paginator = Paginator(conversas, 50)  # 50 por página
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)

    # Buscar códigos de clientes da página atual
    codigos_clientes = [c.codigo for c in page_obj if c.codigo]

    # Buscar TODOS os clientes necessários de uma vez (1 query apenas)
    clientes_dict = {}
    if codigos_clientes:
        clientes = Cliente.objects.filter(codigo__in=codigos_clientes)
        clientes_dict = {str(c.codigo): c for c in clientes}

    # Montar lista com clientes
    conversas_com_cliente = []
    for conversa in page_obj:
        cliente = None
        codigo_vendedor = None

        if conversa.codigo:
            cliente = clientes_dict.get(str(conversa.codigo))
            codigo_vendedor = cliente.codigo_vendedor if cliente else None

        conversas_com_cliente.append({
            'conversa': conversa,
            'cliente': cliente,
            'codigo_vendedor': codigo_vendedor
        })

    # Substituir items do page_obj
    page_obj.object_list = conversas_com_cliente

    context = {
        'page_obj': page_obj,
        'total_conversas': total_conversas,
        'busca': busca,
    }

    return render(request, 'gestor/auditoria_contatos_list.html', context)


@login_required
def auditoria_contatos_export(request):
    """Exporta conversas do Chatwoot para Excel"""

    # Aplicar busca
    busca = request.GET.get('q', '')

    conversas = ChatwootLastConversation.objects.all().order_by('-last_activity')

    if busca:
        conversas = conversas.filter(
            Q(name__icontains=busca) |
            Q(email__icontains=busca) |
            Q(phone__icontains=busca) |
            Q(codigo__icontains=busca) |
            Q(tipo__icontains=busca)
        )

    # Criar workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Auditoria Contatos"

    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    # Cabeçalhos
    headers = [
        'ID Contato', 'Nome', 'Email', 'Telefone',
        'Código Cliente', 'Nome Cliente', 'Código Vendedor', 'Tipo',
        'Última Atividade'
    ]

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Dados
    row = 2
    for conversa in conversas:
        # Buscar cliente e vendedor
        cliente = None
        codigo_vendedor = ''
        if conversa.codigo:
            try:
                cliente = Cliente.objects.get(codigo=conversa.codigo)
                codigo_vendedor = cliente.codigo_vendedor or ''
            except Cliente.DoesNotExist:
                pass

        ws.cell(row=row, column=1, value=conversa.contact_id)
        ws.cell(row=row, column=2, value=conversa.name or '')
        ws.cell(row=row, column=3, value=conversa.email or '')
        ws.cell(row=row, column=4, value=conversa.phone or '')
        ws.cell(row=row, column=5, value=conversa.codigo or '')
        ws.cell(row=row, column=6, value=cliente.nome if cliente else '')
        ws.cell(row=row, column=7, value=codigo_vendedor)
        ws.cell(row=row, column=8, value=conversa.tipo or '')
        ws.cell(row=row, column=9, value=conversa.last_activity.strftime('%d/%m/%Y %H:%M') if conversa.last_activity else '')

        row += 1

    # Ajustar largura das colunas
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 30
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 18

    # Preparar resposta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'auditoria_contatos_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename={filename}'

    wb.save(response)
    return response
