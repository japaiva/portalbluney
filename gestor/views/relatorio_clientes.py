# gestor/views/relatorio_clientes.py

import logging
from datetime import datetime, date
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.db.models import Sum, Q
from django.http import HttpResponse
from django.utils import timezone
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict, OrderedDict
import json

# Importando locale para formatação de data em português
import locale
try:
    locale.setlocale(locale.LC_ALL, 'pt_BR.UTF-8')
except locale.Error:
    try:
        locale.setlocale(locale.LC_ALL, 'Portuguese_Brazil.1252') # Para Windows
    except locale.Error:
        logging.warning("Não foi possível configurar o locale para pt_BR. Os nomes dos meses podem não estar em português.")


from core.models import Cliente, Vendas, Loja, Vendedor, GrupoProduto, Fabricante, Produto

logger = logging.getLogger(__name__)


@login_required
def relatorio_clientes(request):
    """Relatório de clientes com faturamento mensal"""
    
    # ===== OBTER FILTROS (MÚLTIPLA ESCOLHA) =====
    data_inicio = request.GET.get('data_inicio', '')
    data_fim = request.GET.get('data_fim', '')
    
    # Filtros múltiplos - usar getlist() para arrays
    loja_codigos = request.GET.getlist('loja')
    vendedor_codigos = request.GET.getlist('vendedor')
    estados = request.GET.getlist('estado')
    grupo_codigos = request.GET.getlist('grupo')
    fabricante_codigos = request.GET.getlist('fabricante')
    produto_codigos = request.GET.getlist('produto')
    
    incluir_coligados = request.GET.get('incluir_coligados', '') == 'on'
    # Força 'apenas_com_vendas' a ser True, independentemente do input do usuário
    apenas_com_vendas = True 
    
    # ===== DATAS PADRÃO (ÚLTIMOS 12 MESES) =====
    if not data_inicio or not data_fim:
        hoje = date.today()
        primeiro_dia_ano = date(hoje.year, 1, 1)
        data_inicio = primeiro_dia_ano.strftime('%Y-%m-%d')
        data_fim = hoje.strftime('%Y-%m-%d')
    
    # ===== BUSCAR DADOS PARA OS SELECTS =====
    lojas = Loja.objects.filter(ativo=True).order_by('codigo')
    vendedores = Vendedor.objects.filter(ativo=True).order_by('codigo')
    grupos = GrupoProduto.objects.filter(ativo=True).order_by('codigo')
    fabricantes = Fabricante.objects.filter(ativo=True).order_by('codigo')
    produtos = Produto.objects.filter(ativo=True).order_by('codigo')
    
    # Estados únicos do cadastro de clientes
    estados_disponiveis = Cliente.objects.filter(estado__isnull=False).values_list('estado', flat=True).distinct().order_by('estado')
    
    # ===== APLICAR FILTROS E GERAR RELATÓRIO =====
    dados_relatorio = []
    meses_periodo = []
    
    # O relatório será gerado se 'gerar_relatorio' for acionado ou se houver qualquer filtro aplicado
    # removido o "apenas_com_vendas" da condição, pois agora é sempre True
    if request.GET.get('gerar_relatorio') or any(request.GET.getlist(key) for key in ['loja', 'vendedor', 'estado', 'grupo', 'fabricante', 'produto']) or incluir_coligados:
        try:
            dados_relatorio, meses_periodo = gerar_dados_relatorio(
                data_inicio=data_inicio,
                data_fim=data_fim,
                loja_codigos=loja_codigos,
                vendedor_codigos=vendedor_codigos,
                estados=estados,
                grupo_codigos=grupo_codigos,
                fabricante_codigos=fabricante_codigos,
                produto_codigos=produto_codigos,
                incluir_coligados=incluir_coligados,
                apenas_com_vendas=apenas_com_vendas # agora é sempre True
            )
        except Exception as e:
            logger.error(f"Erro ao gerar relatório: {str(e)}")
            dados_relatorio = []
            meses_periodo = []
    
    # ===== EXPORTAR PARA EXCEL =====
    if request.GET.get('exportar_excel') and dados_relatorio:
        return exportar_relatorio_excel(dados_relatorio, meses_periodo, {
            'data_inicio': data_inicio,
            'data_fim': data_fim,
            'loja': ', '.join(loja_codigos) if loja_codigos else 'Todas',
            'vendedor': ', '.join(vendedor_codigos) if vendedor_codigos else 'Todos',
            'estado': ', '.join(estados) if estados else 'Todos',
            'grupo': ', '.join(grupo_codigos) if grupo_codigos else 'Todos',
            'fabricante': ', '.join(fabricante_codigos) if fabricante_codigos else 'Todos',
            'produto': ', '.join(produto_codigos) if produto_codigos else 'Todos',
            'incluir_coligados': incluir_coligados,
            'apenas_com_vendas': apenas_com_vendas
        })
    
    context = {
        # Dados do relatório
        'dados_relatorio': dados_relatorio,
        'meses_periodo': meses_periodo,
        
        # Dados para os selects
        'lojas': lojas,
        'vendedores': vendedores,
        'grupos': grupos,
        'fabricantes': fabricantes,
        'produtos': produtos,
        'estados': estados_disponiveis,
        
        # Filtros aplicados
        'filtros': {
            'data_inicio': data_inicio,
            'data_fim': data_fim,
            'loja_list': loja_codigos,
            'vendedor_list': vendedor_codigos,
            'estado_list': estados,
            'grupo_list': grupo_codigos,
            'fabricante_list': fabricante_codigos,
            'produto_list': produto_codigos,
            'incluir_coligados': incluir_coligados,
            'apenas_com_vendas': apenas_com_vendas, # agora é sempre True
        },
        
        # Estatísticas
        'total_clientes': len(dados_relatorio),
        'total_geral': sum(cliente['total'] for cliente in dados_relatorio) if dados_relatorio else 0,
        
        # JSON para JavaScript (serialização segura)
        'dados_relatorio_json': json.dumps(dados_relatorio) if dados_relatorio else '[]',
        'meses_periodo_json': json.dumps(meses_periodo) if meses_periodo else '[]',
    }
    
    return render(request, 'gestor/relatorio_clientes.html', context)


def gerar_dados_relatorio(data_inicio, data_fim, loja_codigos, vendedor_codigos, estados, 
                         grupo_codigos, fabricante_codigos, produto_codigos, incluir_coligados, apenas_com_vendas):
    """Gera os dados do relatório de clientes com faturamento mensal - MÚLTIPLA ESCOLHA"""
    
    # ===== QUERY BASE DE VENDAS =====
    vendas_query = Vendas.objects.filter(
        data_venda__gte=data_inicio,
        data_venda__lte=data_fim
    ).select_related('cliente', 'loja', 'produto', 'grupo_produto', 'fabricante').exclude(
        cliente__status='outros'
    )
    
    # ===== APLICAR FILTROS NAS VENDAS (MÚLTIPLA ESCOLHA) =====
    if loja_codigos:
        vendas_query = vendas_query.filter(loja__codigo__in=loja_codigos)
    
    if vendedor_codigos:
        vendas_query = vendas_query.filter(vendedor_nf__in=vendedor_codigos)
    
    if grupo_codigos:
        vendas_query = vendas_query.filter(grupo_produto__codigo__in=grupo_codigos)
    
    if fabricante_codigos:
        vendas_query = vendas_query.filter(fabricante__codigo__in=fabricante_codigos)
    
    if produto_codigos:
        vendas_query = vendas_query.filter(produto__codigo__in=produto_codigos)
    
    # ===== FILTRO POR ESTADOS DO CLIENTE (MÚLTIPLA ESCOLHA) =====
    if estados:
        vendas_query = vendas_query.filter(cliente__estado__in=estados)
    
    # ===== GERAR LISTA DE MESES DO PERÍODO =====
    data_inicio_obj = datetime.strptime(data_inicio, '%Y-%m-%d').date()
    data_fim_obj = datetime.strptime(data_fim, '%Y-%m-%d').date()
    
    meses_periodo = []
    ano_mes_atual = data_inicio_obj.replace(day=1)
    
    while ano_mes_atual <= data_fim_obj:
        # Usar %b para o nome do mês abreviado em português (depende do locale)
        # Fallback para nome completo caso o locale não funcione como esperado
        try:
            nome_mes_abreviado = ano_mes_atual.strftime('%b').capitalize()
        except Exception:
            nome_mes_abreviado = ano_mes_atual.strftime('%B')[0:3].capitalize()
        
        meses_periodo.append({
            'ano_mes': ano_mes_atual.strftime('%Y-%m'),
            'nome': f"{nome_mes_abreviado}/{ano_mes_atual.year}", # Formatado como "Jan/2023"
            'ano': ano_mes_atual.year,
            'mes': ano_mes_atual.month
        })
        
        # Próximo mês
        if ano_mes_atual.month == 12:
            ano_mes_atual = ano_mes_atual.replace(year=ano_mes_atual.year + 1, month=1)
        else:
            ano_mes_atual = ano_mes_atual.replace(month=ano_mes_atual.month + 1)
    
    # ===== AGRUPAR VENDAS POR CLIENTE E MÊS =====
    vendas_por_cliente_mes = defaultdict(lambda: defaultdict(float))
    clientes_info = {}
    
    for venda in vendas_query:
        cliente_codigo = venda.cliente.codigo
        ano_mes = venda.data_venda.strftime('%Y-%m')
        
        # Somar valor da venda
        vendas_por_cliente_mes[cliente_codigo][ano_mes] += float(venda.valor_total or 0)
        
        # Armazenar informações do cliente
        if cliente_codigo not in clientes_info:
            clientes_info[cliente_codigo] = {
                'cliente': venda.cliente,
                'nome': venda.cliente.nome,
                'codigo': venda.cliente.codigo,
                'cpf_cnpj': venda.cliente.cpf_cnpj or '-',
                'cidade': venda.cliente.cidade or '-',
                'estado': venda.cliente.estado or '-',
                'vendedor_codigo': venda.cliente.codigo_vendedor or '-',
                'vendedor_nome': venda.cliente.nome_vendedor or '-',
                'loja_codigo': venda.cliente.codigo_loja or '-',
                'status': venda.cliente.get_status_display(),
                'tipo': 'Coligado' if venda.cliente.codigo_master else 'Principal',
            }
    
    # ===== INCLUIR CLIENTES SEM VENDAS (SE SOLICITADO) =====
    # Apenas com vendas agora é sempre True no contexto, então o bloco abaixo só é executado
    # se apenas_com_vendas for False, o que nunca acontecerá.
    # No entanto, mantemos a estrutura caso a regra de negócio mude no futuro.
    if not apenas_com_vendas:
        clientes_sem_vendas = Cliente.objects.exclude(status='outros')
        
        # Aplicar filtros de cliente (MÚLTIPLA ESCOLHA)
        if estados:
            clientes_sem_vendas = clientes_sem_vendas.filter(estado__in=estados)
        
        if vendedor_codigos:
            clientes_sem_vendas = clientes_sem_vendas.filter(codigo_vendedor__in=vendedor_codigos)
        
        if loja_codigos:
            clientes_sem_vendas = clientes_sem_vendas.filter(codigo_loja__in=loja_codigos)
        
        # Filtro de coligados
        if not incluir_coligados:
            clientes_sem_vendas = clientes_sem_vendas.filter(
                Q(codigo_master__isnull=True) | Q(codigo_master='')
            )
        
        # Adicionar clientes sem vendas
        for cliente in clientes_sem_vendas:
            if cliente.codigo not in clientes_info:
                clientes_info[cliente.codigo] = {
                    'cliente': cliente,
                    'nome': cliente.nome,
                    'codigo': cliente.codigo,
                    'cpf_cnpj': cliente.cpf_cnpj or '-',
                    'cidade': cliente.cidade or '-',
                    'estado': cliente.estado or '-',
                    'vendedor_codigo': cliente.codigo_vendedor or '-',
                    'vendedor_nome': cliente.nome_vendedor or '-',
                    'loja_codigo': cliente.codigo_loja or '-',
                    'status': cliente.get_status_display(),
                    'tipo': 'Coligado' if cliente.codigo_master else 'Principal',
                }
    
    # ===== APLICAR FILTRO DE COLIGADOS =====
    if not incluir_coligados:
        clientes_filtrados = {}
        for codigo, info in clientes_info.items():
            if info['tipo'] == 'Principal':
                clientes_filtrados[codigo] = info
        clientes_info = clientes_filtrados
    
    # ===== MONTAR DADOS FINAIS =====
    dados_relatorio = []
    
    for cliente_codigo, cliente_info in clientes_info.items():
        # Dados do cliente
        linha_cliente = {
            'codigo': cliente_info['codigo'],
            'nome': cliente_info['nome'],
            'cpf_cnpj': cliente_info['cpf_cnpj'],
            'cidade': cliente_info['cidade'],
            'estado': cliente_info['estado'],
            'vendedor_codigo': cliente_info['vendedor_codigo'],
            'vendedor_nome': cliente_info['vendedor_nome'],
            'loja_codigo': cliente_info['loja_codigo'],
            'status': cliente_info['status'],
            'tipo': cliente_info['tipo'],
            'vendas_por_mes': {},
            'total': 0
        }
        
        # Valores por mês
        total_cliente = 0
        for mes_info in meses_periodo:
            ano_mes = mes_info['ano_mes']
            valor_mes = vendas_por_cliente_mes[cliente_codigo].get(ano_mes, 0)
            linha_cliente['vendas_por_mes'][ano_mes] = valor_mes
            total_cliente += valor_mes
        
        linha_cliente['total'] = total_cliente
        dados_relatorio.append(linha_cliente)
    
    # ===== ORDENAR POR TOTAL DECRESCENTE =====
    dados_relatorio.sort(key=lambda x: x['total'], reverse=True)
    
    return dados_relatorio, meses_periodo


def exportar_relatorio_excel(dados_relatorio, meses_periodo, filtros):
    """Exporta o relatório para Excel"""
    
    # ===== CRIAR WORKBOOK =====
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Relatório de Clientes"
    
    # ===== ESTILOS =====
    font_titulo = Font(name='Arial', size=14, bold=True)
    font_cabecalho = Font(name='Arial', size=10, bold=True, color='FFFFFF')
    font_dados = Font(name='Arial', size=9)
    font_total = Font(name='Arial', size=10, bold=True)
    
    fill_cabecalho = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    fill_total = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
    
    border_thin = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    alignment_center = Alignment(horizontal='center', vertical='center')
    alignment_right = Alignment(horizontal='right', vertical='center')
    
    # ===== TÍTULO =====
    ws.merge_cells('A1:' + get_column_letter(11 + len(meses_periodo)) + '1')
    titulo_cell = ws['A1']
    titulo_cell.value = "RELATÓRIO DE CLIENTES - FATURAMENTO MENSAL"
    titulo_cell.font = font_titulo
    titulo_cell.alignment = alignment_center
    
    # ===== INFORMAÇÕES DOS FILTROS =====
    linha_atual = 3
    ws[f'A{linha_atual}'] = "Período:"
    ws[f'B{linha_atual}'] = f"{filtros['data_inicio']} a {filtros['data_fim']}"
    linha_atual += 1
    
    if filtros['loja'] and filtros['loja'] != 'Todas': # Verificando se é diferente de 'Todas'
        ws[f'A{linha_atual}'] = "Loja:"
        ws[f'B{linha_atual}'] = filtros['loja']
        linha_atual += 1
    
    if filtros['vendedor'] and filtros['vendedor'] != 'Todos': # Verificando se é diferente de 'Todos'
        ws[f'A{linha_atual}'] = "Vendedor:"
        ws[f'B{linha_atual}'] = filtros['vendedor']
        linha_atual += 1
    
    if filtros['estado'] and filtros['estado'] != 'Todos': # Verificando se é diferente de 'Todos'
        ws[f'A{linha_atual}'] = "Estado:"
        ws[f'B{linha_atual}'] = filtros['estado']
        linha_atual += 1
    
    if filtros['grupo'] and filtros['grupo'] != 'Todos': # Verificando se é diferente de 'Todos'
        ws[f'A{linha_atual}'] = "Grupo:"
        ws[f'B{linha_atual}'] = filtros['grupo']
        linha_atual += 1

    if filtros['fabricante'] and filtros['fabricante'] != 'Todos': # Verificando se é diferente de 'Todos'
        ws[f'A{linha_atual}'] = "Fabricante:"
        ws[f'B{linha_atual}'] = filtros['fabricante']
        linha_atual += 1

    if filtros['produto'] and filtros['produto'] != 'Todos': # Verificando se é diferente de 'Todos'
        ws[f'A{linha_atual}'] = "Produto:"
        ws[f'B{linha_atual}'] = filtros['produto']
        linha_atual += 1
    
    ws[f'A{linha_atual}'] = "Incluir Coligados:"
    ws[f'B{linha_atual}'] = "Sim" if filtros['incluir_coligados'] else "Não"
    linha_atual += 1

    ws[f'A{linha_atual}'] = "Apenas com Vendas:"
    ws[f'B{linha_atual}'] = "Sim" if filtros['apenas_com_vendas'] else "Não"
    linha_atual += 1
    
    linha_atual += 1
    
    # ===== CABEÇALHOS =====
    cabecalhos = [
        'Código', 'Nome do Cliente', 'CPF/CNPJ', 'Cidade', 'UF', 
        'Loja', 'Vendedor', 'Status', 'Tipo'
    ]
    
    # Adicionar meses
    for mes_info in meses_periodo:
        cabecalhos.append(mes_info['nome'])
    
    cabecalhos.append('TOTAL')
    
    # Escrever cabeçalhos
    for col, cabecalho in enumerate(cabecalhos, 1):
        cell = ws.cell(row=linha_atual, column=col)
        cell.value = cabecalho
        cell.font = font_cabecalho
        cell.fill = fill_cabecalho
        cell.alignment = alignment_center
        cell.border = border_thin
    
    linha_atual += 1
    
    # ===== DADOS =====
    total_geral_por_mes = {}
    total_geral = 0
    
    for cliente in dados_relatorio:
        col = 1
        
        # Dados básicos do cliente
        ws.cell(row=linha_atual, column=col, value=cliente['codigo']).font = font_dados
        col += 1
        ws.cell(row=linha_atual, column=col, value=cliente['nome']).font = font_dados
        col += 1
        ws.cell(row=linha_atual, column=col, value=cliente['cpf_cnpj']).font = font_dados
        col += 1
        ws.cell(row=linha_atual, column=col, value=cliente['cidade']).font = font_dados
        col += 1
        ws.cell(row=linha_atual, column=col, value=cliente['estado']).font = font_dados
        col += 1
        ws.cell(row=linha_atual, column=col, value=cliente['loja_codigo']).font = font_dados
        col += 1
        ws.cell(row=linha_atual, column=col, value=f"{cliente['vendedor_codigo']} - {cliente['vendedor_nome']}" if cliente['vendedor_codigo'] != '-' else '-').font = font_dados
        col += 1
        ws.cell(row=linha_atual, column=col, value=cliente['status']).font = font_dados
        col += 1
        ws.cell(row=linha_atual, column=col, value=cliente['tipo']).font = font_dados
        col += 1
        
        # Valores por mês
        for mes_info in meses_periodo:
            ano_mes = mes_info['ano_mes']
            valor = cliente['vendas_por_mes'].get(ano_mes, 0)
            
            cell = ws.cell(row=linha_atual, column=col)
            if valor > 0:
                cell.value = valor
                cell.number_format = '#,##0.00'
            else:
                cell.value = '-'
            cell.font = font_dados
            cell.alignment = alignment_right
            
            # Somar para total geral
            if ano_mes not in total_geral_por_mes:
                total_geral_por_mes[ano_mes] = 0
            total_geral_por_mes[ano_mes] += valor
            
            col += 1
        
        # Total do cliente
        cell = ws.cell(row=linha_atual, column=col)
        if cliente['total'] > 0:
            cell.value = cliente['total']
            cell.number_format = '#,##0.00'
        else:
            cell.value = '-'
        cell.font = font_total
        cell.alignment = alignment_right
        cell.fill = fill_total
        
        total_geral += cliente['total']
        linha_atual += 1
    
    # ===== LINHA DE TOTAIS =====
    linha_total = linha_atual
    col = 1
    
    # Células vazias até "TOTAL"
    for i in range(8):
        ws.cell(row=linha_total, column=col).fill = fill_total
        col += 1
    
    # Label "TOTAL"
    cell = ws.cell(row=linha_total, column=col)
    cell.value = "TOTAL GERAL"
    cell.font = font_total
    cell.fill = fill_total
    cell.alignment = alignment_right
    col += 1
    
    # Totais por mês
    for mes_info in meses_periodo:
        ano_mes = mes_info['ano_mes']
        cell = ws.cell(row=linha_total, column=col)
        valor_total_mes = total_geral_por_mes.get(ano_mes, 0)
        if valor_total_mes > 0:
            cell.value = valor_total_mes
            cell.number_format = '#,##0.00'
        else:
            cell.value = '-'
        cell.font = font_total
        cell.fill = fill_total
        cell.alignment = alignment_right
        col += 1
    
    # Total geral
    cell = ws.cell(row=linha_total, column=col)
    if total_geral > 0:
        cell.value = total_geral
        cell.number_format = '#,##0.00'
    else:
        cell.value = '-'
    cell.font = font_total
    cell.fill = fill_total
    cell.alignment = alignment_right
    
    # ===== BORDAS PARA TODA A TABELA =====
    for row in range(linha_atual - len(dados_relatorio), linha_atual + 1):
        for col in range(1, len(cabecalhos) + 1):
            ws.cell(row=row, column=col).border = border_thin
    
    # ===== AJUSTAR LARGURAS DAS COLUNAS =====
    larguras = [12, 30, 18, 20, 5, 8, 25, 12, 12]  # Colunas básicas
    for mes in meses_periodo:
        larguras.append(12)  # Colunas de meses
    larguras.append(15)  # Coluna de total
    
    for i, largura in enumerate(larguras, 1):
        ws.column_dimensions[get_column_letter(i)].width = largura
    
    # ===== PREPARAR RESPOSTA HTTP =====
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    # Nome do arquivo
    data_hoje = timezone.now().strftime('%Y%m%d_%H%M%S')
    response['Content-Disposition'] = f'attachment; filename="relatorio_clientes_{data_hoje}.xlsx"'
    
    # Salvar workbook na resposta
    wb.save(response)
    return response